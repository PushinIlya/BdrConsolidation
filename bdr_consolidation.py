#!/usr/bin/env python
# coding: utf-8

# In[ ]:


'''
Скрипт для консолидации выгруженных из 1С БДР и их распределения по папкам 
в соответствии с периодом, по которому они сформированы
'''
# импортируем библиотеки и модули
import numpy as np
import pandas as pd
import os
import math
import shutil
import copy
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = os.path.join('C:',
                      'Исходные данные')

# создадим список с названиями всех файлов в папке с исходными данными с расширением .xlsx
raw_data_name_list = [raw_data for raw_data in os.listdir(source) if raw_data.startswith('БДР') and
                      raw_data.endswith('.xlsx')]

print(f'Исходные данные: {raw_data_name_list}')


# In[ ]:


print('Выполнение скрипта "bdr_consolidation" (консолидация БДР)')
print(f'Чтение БДР из "{source}"')

# создадим справочник из всех датафреймов в списке 'raw_data_name_list'
all_raw_data = {raw_data:pd.read_excel(os.path.join(source, f'{raw_data}')) for raw_data in tqdm(raw_data_name_list)}


# In[ ]:


# создадим список столбцов, которые нужно сохранить
columns = [0,1,-8,-7,-6,-5,-4,-3,-2,-1]


# In[ ]:


# удалим столбцы, содержащие только пустые значения
all_raw_data = {raw_data[0]: raw_data[1].dropna(axis=1, how='all', inplace=False) for raw_data in all_raw_data.items()}


# In[ ]:


# оставим только столбцы из списка 'columns'
all_raw_data = {raw_data[0]: raw_data[1].iloc[:, columns] for raw_data in all_raw_data.items()}


# In[ ]:


# создадим список 'source_file_path' и заполним его значениями путей к исходным файлам
source_file_path = []
for name in raw_data_name_list:
    source_file_path.append(os.path.join(source, name))


# In[ ]:


# объединим списки 'raw_data_name_list' и 'source_file_path'
# в список кортежей с названием файла и путём к нему
source_file = list(zip(raw_data_name_list, source_file_path))


# In[ ]:


'''
Цикл, который перезаписывает каждый файл из 'all_raw_data', чтобы сохранить внесённые ранее изменения в его структуру,
а также сохраняет файл в формате xlsx в первоначальную папку
'''
print(f'Перезапись БДР из "{source}"')
for raw_data in tqdm(all_raw_data.items()):
    for file in source_file:
        if raw_data[0] == file[0]:
            raw_data[1].to_excel(file[1], header=False, index=False)


# In[ ]:


# создадим список с периодами формирования БДР для каждого файла
period_list = []
for raw_data in all_raw_data.values():
    period_list.append(raw_data.iloc[2,2])


# In[ ]:


# переименуем каждый датафрейм, оставив только название компании в имени
all_raw_data = {raw_data[0][raw_data[0].rfind('_')+1:]: raw_data[1] for raw_data in all_raw_data.items()}


# In[ ]:


# создадим excel файл 'Сводный_БДР' и добавим в него все датафреймы с исходными данными
with pd.ExcelWriter(os.path.join(source, f'Сводный_БДР_{period_list[0]}.xlsx')) as source_file:
    print(f'Создание листов excel из БДР')
    for raw_data in tqdm(all_raw_data.items()):
        raw_data[1].to_excel(source_file,
                             sheet_name=raw_data[0][raw_data[0].rfind('_')+1:].replace('.xlsx', ''),
                             header=False,
                             index=False)


# In[ ]:


# функция для извлечения номера месяца по названию месяца
def month_number(month_name):
    months = {'январь': '01',
              'февраль': '02', 
              'март': '03',
              'апрель': '04', 
              'май': '05', 
              'июнь': '06',
              'июль': '07',
              'август': '08',
              'сентябрь': '09',
              'октябрь': '10', 
              'ноябрь': '11', 
              'декабрь': '12'}
    month_number = months[month_name]
    return month_number


# In[ ]:


# удалим лишние строки в датафреймах
all_raw_data = {raw_data[0]: raw_data[1][5:] for raw_data in all_raw_data.items()}


# In[ ]:


# переименуем столбцы в каждом датафрейме
for raw_data in all_raw_data.values():
    raw_data.columns = list(range(0,10))


# In[ ]:


# словарь с именами столбцов и типами данных
type_of_columns = {2: 'float64', 
                   3: 'float64', 
                   4: 'float64',
                   5: 'float64', 
                   6: 'float64', 
                   7: 'float64', 
                   8: 'float64',
                   9: 'float64'}


# In[ ]:


# изменим тип данных в каждом датафрейме в соответствии со словарём 'type_of_columns'
all_raw_data = {raw_data[0]: raw_data[1].astype(type_of_columns) for raw_data in all_raw_data.items()}


# In[ ]:


# заменим пустое значение кодификатора в столбце 1 в строке 'Итого', чтобы при конкатенации сохранить строку 'Итого'
for raw_data in all_raw_data.values():
    raw_data[1] = raw_data[1].fillna('999')


# In[ ]:


# объединим данные по каждому датафрейму в единый датафрейм, путём суммирования построчно
consolidated_data = (pd.concat(all_raw_data.values())
                     .groupby([0,1])
                     .sum()
                     .sort_values([1])
                     .reset_index()
                    )


# In[ ]:


# избавимся от лишнего кодификатора в объединённом датафрейме
consolidated_data[1] = consolidated_data[1].replace('999', '')


# In[ ]:


# избавимся от лишнего кодификатора в каждом исходном файле
for raw_data in all_raw_data.values():
    raw_data[1] = raw_data[1].replace('999', '')


# In[ ]:


# рассчитаем значения в столбце 'Отклонение (отн., %)' за месяц
# рассчитаем значения в столбце 'Отклонение (отн., %)' за весь период
consolidated_data[5] = consolidated_data[4]/abs(consolidated_data[2])*100
consolidated_data[9] = consolidated_data[8]/abs(consolidated_data[6])*100


# In[ ]:


# при расчёте значений 'Отклонение (отн., %)' при делении на 0 получились бесконечные числа, заменим их на 100
consolidated_data.replace([np.inf, -np.inf], 100, inplace=True)


# In[ ]:


# переименуем столбцы
consolidated_data.columns = ['Статья бюджета',
                             'Кодификатор',
                             f'План годовой по месяцам ({period_list[0]})',
                             f'ФАКТ ({period_list[0]})',
                             f'Отклонение (абс.) ({period_list[0]})',
                             f'Отклонение (отн., %) ({period_list[0]})',
                             'План годовой по месяцам (Итого)',
                             'ФАКТ (Итого)',
                             'Отклонение (абс.) (Итого)',
                             'Отклонение (отн., %) (Итого)']


# In[ ]:


book_path = os.path.join(source, f'Сводный_БДР_{period_list[0]}.xlsx') # создадим путь к итоговому excel файлу
book = load_workbook(book_path) # загрузим excel файл для редактирования


# In[ ]:


columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'] # создадим список столбцов, к которым будем обращаться


# In[ ]:


# заменим названия столбцов на каждом листе
for sheet in book:
    x = 0
    for column in columns:
        sheet[f'{column}5'] = consolidated_data.columns[x]
        x += 1


# In[ ]:


# создадим новый лист
new_sheet = book.create_sheet('Консолидация')

# загрузим объединённый датафрейм на новый лист
for row in dataframe_to_rows(consolidated_data, index=False, header=True):
    new_sheet.append(row)


# In[ ]:


# создадим таблицы на каждом листе
for sheet in book:
    if sheet.title == 'Консолидация':
        table = Table(displayName=sheet.title, ref=f'A1:J{sheet.max_row}') # создадим таблицу
        style = TableStyleInfo(name='TableStyleLight13', # создадим стиль таблицы
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)
        table.tableStyleInfo = style # применим стиль
        sheet.add_table(table) # добавим таблицу на лист
    else:
        table = Table(displayName=sheet.title, ref=f'A5:J{sheet.max_row}')
        style = TableStyleInfo(name='TableStyleLight13',
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)


# In[ ]:


# функция для группировки строк на каждом листе, кроме 'Консолидация' в excel файле на основе столбца 'Кодификатор'
def rows_grouping(lvl):
    # группировка 1-го уровня
    if lvl == 1:
        # создадим список с индексами строк для группировки 1-го уровня
        rows_start = {raw_data[0]: list(raw_data[1][1]
                                        .str.count('\\.')[lambda x: x == 1]
                                        .index) 
                      for raw_data in all_raw_data.items()}
        
        # добавим в список с индексами строк для группировки 1-го уровня последнюю строку на листе,
        # чтобы группировка была корректной 
        for raw_data in all_raw_data.items():
            rows_start[raw_data[0]].append(list(raw_data[1][1].index)[-1])
            
        # создадим аналогичный список с индексами строк для группировки 1-го уровня,
        # только смещённый на 1-ну позицию
        rows_end = {raw_data[0]: rows_start[raw_data[0]][1:]
                    for raw_data in all_raw_data.items()}
        # пересоздадим список с индексами строк для группировки 1-го уровня, так как до этого его изменили
        rows_start = {raw_data[0]: list(raw_data[1][1]
                                        .str.count('\\.')[lambda x: x == 1]
                                        .index)
                      for raw_data in all_raw_data.items()}
        # объединим оба списка в список кортежей
        rows = {raw_data[0]: list(zip(rows_start[raw_data[0]], rows_end[raw_data[0]]))
                for raw_data in all_raw_data.items()}
    
        return rows
    else:
        temp_all_raw_data = copy.deepcopy(all_raw_data) # создадим копию 'all_raw_data'
        for raw_data in temp_all_raw_data.values():
            for row in raw_data[1]:
                if not (pd.isna(row)): # не учитываем пустые значения
                    # группировка 2-го уровня
                    if lvl == 2:
                        # оставим только значения с 2-мя точками в столбце с кодификатором
                        raw_data.replace({row: row[:row.find('.', row.find('.')+1)+1]}, inplace=True) 
                    # группировка 3-го уровня
                    elif lvl == 3:
                        # оставим только значения с 3-мя точками в столбце с кодификатором
                        raw_data.replace({row: row[:row.find('.', row.find('.', row.find('.')+1)+1)+1]}, inplace=True)

        # создадим список с индексами строк для группировки, оставив только первое вхождение кодификатора
        # с необходимым уровнем группировки
        rows_start = {raw_data[0]: list(raw_data[1][1][lambda x: x.str.count('\\.') == lvl]
                                        .drop_duplicates(keep='first')
                                        .index)
                      for raw_data in temp_all_raw_data.items()}
        # создадим список с индексами строк для группировки, оставив только последнее вхождение кодификатора
        # с необходимым уровнем группировки
        rows_end = {raw_data[0]: list(raw_data[1][1][lambda x: x.str.count('\\.') == lvl]
                                      .drop_duplicates(keep='last')
                                      .index)
                    for raw_data in temp_all_raw_data.items()}
        # объединим оба списка в список кортежей
        rows = {raw_data[0]: list(zip(rows_start[raw_data[0]], rows_end[raw_data[0]])) 
                for raw_data in temp_all_raw_data.items()}
    
        return rows


# In[ ]:


# функция для группировки строк в excel файле на основе столбца 'Кодификатор' на листе 'Консолидация'
# код аналогичен функции 'rows_grouping'
def rows_grouping_conso(lvl):
    if lvl == 1:
        rows_start = list(consolidated_data['Кодификатор'].str.count('\\.')[lambda x: x == 1].index)
        rows_start.append(list(consolidated_data['Кодификатор'].index)[-1])
    
        rows_end = rows_start[1:]
        rows_start = list(consolidated_data['Кодификатор'].str.count('\\.')[lambda x: x == 1].index)
        rows = list(zip(rows_start, rows_end))
    
        return rows
    else:
        temp_consolidated_data = copy.deepcopy(consolidated_data)
        for row in temp_consolidated_data['Кодификатор']:
            if not (pd.isna(row)):
                if lvl == 2:
                    temp_consolidated_data['Кодификатор'].replace({row: row[:row.find('.',
                                                                                      row.find('.')+1)+1]},
                                                                  inplace=True)
                elif lvl == 3:
                    temp_consolidated_data['Кодификатор'].replace({row: row[:row.find('.',
                                                                                      row.find('.',
                                                                                               row.find('.')+1)+1)+1]},
                                                                  inplace=True)

        rows_start = list(temp_consolidated_data['Кодификатор'][lambda x: x.str.count('\\.') == lvl]
                          .drop_duplicates(keep='first')
                          .index)
        rows_end = list(temp_consolidated_data['Кодификатор'][lambda x: x.str.count('\\.') == lvl]
                        .drop_duplicates(keep='last')
                        .index)
        rows = list(zip(rows_start, rows_end))
    
        return rows


# In[ ]:


# функция для применения группировки строк на каждом листе, кроме 'Консолидация' в excel файле
def apply_rows_grouping():
    for row_key, row_value in rows_1.items():
        for row in row_value:
            if (row[1] - row[0]) > 0:
                '''
                К каждой начальной строке добавляем 2 для корректной группировки,
                добавили 1, так как нумерация в excel начинается с 1-ой строки, а в python с 0
                и ещё 1, чтобы начальная строка не скрылась из-за группировки.
                К каждой конечной строке ничего не добавляем, так как
                весь диапазон группировки достигнут.
                ''' 
                book[row_key.replace('.xlsx', '')].row_dimensions.group(row[0]+2,
                                                                        row[1],
                                                                        outline_level=1,
                                                                        hidden=True)
    for row_key, row_value in rows_2.items():
        for row in row_value:
            if (row[1] - row[0]) > 0:
                '''
                К каждой начальной строке добавляем 2 для корректной группировки,
                добавили 1, так как нумерация в excel начинается с 1-ой строки, а в python с 0
                и ещё 1, чтобы начальная строка не скрылась из-за группировки.
                К каждой конечной строке добавляем 1, чтобы достичь весь диапазон
                группировки.
                ''' 
                book[row_key.replace('.xlsx', '')].row_dimensions.group(row[0]+2, 
                                                                        row[1]+1,
                                                                        outline_level=2,
                                                                        hidden=True)
    for row_key, row_value in rows_3.items():
        for row in row_value:
            if (row[1] - row[0]) > 0:
                # аналогично группировке 2-го уровня
                book[row_key.replace('.xlsx', '')].row_dimensions.group(row[0]+2, 
                                                                        row[1]+1,
                                                                        outline_level=3, 
                                                                        hidden=True)


# In[ ]:


'''
Функция для применения группировки строк в excel файле на листе 'Консолидация'.
Код аналогичен функции 'apply_rows_grouping', 
только к каждой строке добавляем на 1 больше, так как 'consolidated_data'
выгружена с заголовками, поэтому нумерацию нужно сместить.
'''
def apply_rows_grouping_conso():
    for row in rows_1_conso:
        if (row[1] - row[0]) > 0:
            book['Консолидация'].row_dimensions.group(row[0]+3,
                                                      row[1]+1,
                                                      outline_level=1,
                                                      hidden=True)
    for row in rows_2_conso:
        if (row[1] - row[0]) > 0:
            book['Консолидация'].row_dimensions.group(row[0]+3, 
                                                      row[1]+2,
                                                      outline_level=2,
                                                      hidden=True)
    for row in rows_3_conso:
        if (row[1] - row[0]) > 0:
            book['Консолидация'].row_dimensions.group(row[0]+3,
                                                      row[1]+2, 
                                                      outline_level=3, 
                                                      hidden=True)


# In[ ]:


# создадим для каждого уровня группировки на каждом листе, кроме 'Консолидация' список кортежей
rows_1 = rows_grouping(1)
rows_2 = rows_grouping(2)
rows_3 = rows_grouping(3)
# создадим для каждого уровня группировки на листе 'Консолидация' список кортежей
rows_1_conso = rows_grouping_conso(1)
rows_2_conso = rows_grouping_conso(2)
rows_3_conso = rows_grouping_conso(3)


# In[ ]:


# на основе списков кортежей сгруппируем строки на каждом листе, кроме 'Консолидация' в excel файле
apply_rows_grouping()
# на основе списков кортежей сгруппируем строки на листе 'Консолидация' в excel файле
apply_rows_grouping_conso()


# In[ ]:


# цикл для изменения ширины столбцов на каждом листе в excel файле
for sheet in book:
    for column in columns:
        sheet.column_dimensions[column].width = 20
    sheet.column_dimensions['A'].width = 65
    sheet.column_dimensions['B'].width = 20


# In[ ]:


# цикл для изменения формата ячеек на каждом листе в excel файле
for sheet in book:
    for column in columns:
        if column in ['A', 'B']:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[0]  
        else:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[3]


# In[ ]:


# цикл для изменения стиля и выравнивания ячеек на каждом листе в excel файле
for sheet in book:
    if sheet.title == 'Консолидация':
        for column in columns:
            # изменение параметров шрифта
            sheet[f'{column}1'].font = Font(bold=True, 
                                            color='FFFFFFFF')
            # изменение выравнивания
            sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                      vertical='center',
                                                      wrap_text=True)
            
            # изменение заливки и шрифта сгруппированных строк
            for row in list(consolidated_data['Кодификатор'][lambda x: x.str.count('\\.') == 1].index):
                pattern_fill = PatternFill(fill_type='solid',
                                           start_color='DAEEF3',
                                           end_color='DAEEF3')
                sheet[f'{column}{row+2}'].fill = pattern_fill # к строке добавляем 2 из-за смещения нумерации
                sheet[f'{column}{row+2}'].font = Font(bold=True)
            for row in list(consolidated_data['Кодификатор'][lambda x: x.str.count('\\.') == 2].index):
                sheet[f'{column}{row+2}'].font = Font(bold=True)
            # изменение заливки и шрифта строки 'Итого'
            sheet[f'{column}{sheet.max_row}'].fill = pattern_fill
            sheet[f'{column}{sheet.max_row}'].font = Font(bold=True)
        
        # изменение выравнивания числовых ячеек
        for cell_tuple in sheet[f'C2:J{sheet.max_row}']:
            for cell in cell_tuple:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
    else:
        for column in columns:
            # изменение параметров шрифта
            sheet[f'{column}5'].font = Font(bold=True, 
                                            color='FFFFFFFF')
            # изменение выравнивания
            sheet[f'{column}5'].alignment = Alignment(horizontal='center',
                                                      vertical='center',
                                                      wrap_text=True)
            
        # изменение заливки и шрифта сгруппированных строк
        for raw_data in all_raw_data.items():
            rows_1 = list(raw_data[1][1].str.count('\\.')[lambda x: x == 1].index)
            rows_2 = list(raw_data[1][1].str.count('\\.')[lambda x: x == 2].index)
            if raw_data[0].replace('.xlsx', '') == sheet.title:
                for column in columns:
                    for row in rows_1:
                        patern_fill = PatternFill(fill_type='solid',
                                                  start_color='DAEEF3',
                                                  end_color='DAEEF3')
                        sheet[f'{column}{row+1}'].fill = patern_fill # к строке добавляем 1 из-за смещения нумерации
                        sheet[f'{column}{row+1}'].font = Font(bold=True)
                    for row in rows_2:
                        sheet[f'{column}{row+1}'].font = Font(bold=True)
                    # изменение заливки и шрифта строки 'Итого'
                    sheet[f'{column}{sheet.max_row}'].fill = patern_fill
                    sheet[f'{column}{sheet.max_row}'].font = Font(bold=True)

        # изменение выравнивания числовых ячеек
        for cell_tuple in sheet[f'C6:J{sheet.max_row}']:
            for cell in cell_tuple:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')


# In[ ]:


# цикл для изменения отступа ячеек на каждом листе в excel файле
for sheet in book:
    if sheet.title == 'Консолидация':
        rows_1 = list(consolidated_data['Кодификатор'].str.count('\\.')[lambda x: x == 2].index)
        for row in rows_1:
            sheet[f'A{row+2}'].alignment = Alignment(indent=1) # к строке добавляем 2 из-за смещения нумерации
            
        rows_2 = list(consolidated_data['Кодификатор'].str.count('\\.')[lambda x: x == 3].index)
        for row in rows_2:
            sheet[f'A{row+2}'].alignment = Alignment(indent=2)
            
        rows_3 = list(consolidated_data['Кодификатор'].str.count('\\.')[lambda x: x == 4].index)
        for row in rows_3:
            sheet[f'A{row+2}'].alignment = Alignment(indent=3)
    else:
        rows_1 = list(all_raw_data[f'{sheet.title}.xlsx'][1].str.count('\\.')[lambda x: x == 2].index)
        for row in rows_1:
            sheet[f'A{row+1}'].alignment = Alignment(indent=1) # к строке добавляем 1 из-за смещения нумерации
            
        rows_2 = list(all_raw_data[f'{sheet.title}.xlsx'][1].str.count('\\.')[lambda x: x == 3].index)
        for row in rows_2:
            sheet[f'A{row+1}'].alignment = Alignment(indent=2)
            
        rows_3 = list(all_raw_data[f'{sheet.title}.xlsx'][1].str.count('\\.')[lambda x: x == 4].index)
        for row in rows_3:
            sheet[f'A{row+1}'].alignment = Alignment(indent=3)


# In[ ]:


book.save(book_path) # сохраним excel файл


# In[ ]:


# создадим список с расположением новых папок
folder_location = os.path.join(source,
                               period_list[0][-4:],
                               f'{month_number(str.lower(period_list[0][:-5]))}_{str.lower(period_list[0][:-5])}')
# создаются только те папки, которые до этого не существовали
if not os.path.exists(folder_location):
    os.makedirs(folder_location)


# In[ ]:


print(f'Распределение БДР и итогового файла из "{source}" по папкам')
# скопируем итоговый файл с расширением .xlsx в новую папку
shutil.copy(book_path, folder_location)
# удалим итоговый файл с расширением .xlsx из первоначальной папки
os.remove(book_path)

# цикл, который копирует каждый файл из 'raw_data_list' в новую папку и удаляет его из папки c исходными данными
for name in tqdm(raw_data_name_list):
    # скопируем исходные данные с расширением .xlsx в новую папку
    shutil.copy(os.path.join(source, name), folder_location)
    # удалим исходные данные с расширением .xlsx из первоначальной папки
    os.remove(os.path.join(source, name))

